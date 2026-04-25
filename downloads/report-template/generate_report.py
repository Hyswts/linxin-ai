#!/usr/bin/env python3
"""
数据自动化报表生成器 v1.1
林昕AI — 企业AI自动化服务

使用方法：
  python generate_report.py                    # 使用内置样例数据
  python generate_report.py your_data.csv      # 使用你的CSV数据

输出：销售数据自动化报表.xlsx
"""

import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.chart import BarChart, PieChart, LineChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.series import DataPoint
from openpyxl.utils import get_column_letter
from datetime import datetime, timedelta
import random
import os
import sys

# ==============================
# 配色系统
# ==============================
COLORS = {
    "green": "22C55E", "green_dark": "15803D", "green_light": "4ADE80",
    "green_bg": "F0FDF4", "white": "FFFFFF", "bg_surface": "F8FAF9",
    "text_primary": "1A2E23", "text_secondary": "4A6B5A", "text_muted": "7CA892",
    "border": "D1E7DD", "orange": "F59E0B", "red": "EF4444",
}

FILL_GREEN = PatternFill(start_color=COLORS["green"], end_color=COLORS["green"], fill_type="solid")
FILL_GREEN_DARK = PatternFill(start_color=COLORS["green_dark"], end_color=COLORS["green_dark"], fill_type="solid")
FILL_GREEN_BG = PatternFill(start_color=COLORS["green_bg"], end_color=COLORS["green_bg"], fill_type="solid")
FILL_WHITE = PatternFill(start_color=COLORS["white"], end_color=COLORS["white"], fill_type="solid")
FILL_SURFACE = PatternFill(start_color=COLORS["bg_surface"], end_color=COLORS["bg_surface"], fill_type="solid")
FILL_RED_LIGHT = PatternFill(start_color="FEF2F2", end_color="FEF2F2", fill_type="solid")

FONT_HEADER = Font(name="微软雅黑", size=12, bold=True, color=COLORS["white"])
FONT_TITLE = Font(name="微软雅黑", size=16, bold=True, color=COLORS["text_primary"])
FONT_BODY = Font(name="微软雅黑", size=10, color=COLORS["text_primary"])
FONT_BODY_BOLD = Font(name="微软雅黑", size=10, bold=True, color=COLORS["text_primary"])
FONT_MUTED = Font(name="微软雅黑", size=10, color=COLORS["text_muted"])
FONT_METRIC = Font(name="微软雅黑", size=24, bold=True, color=COLORS["green_dark"])
FONT_METRIC_LABEL = Font(name="微软雅黑", size=10, color=COLORS["text_secondary"])

ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
ALIGN_LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
ALIGN_RIGHT = Alignment(horizontal="right", vertical="center")

BORDER_THIN = Border(
    left=Side(style="thin", color=COLORS["border"]),
    right=Side(style="thin", color=COLORS["border"]),
    top=Side(style="thin", color=COLORS["border"]),
    bottom=Side(style="thin", color=COLORS["border"]),
)

FMT_CNY = '¥#,##0'      # 货币格式
FMT_PCT = '0.0%'         # 百分比格式
FMT_NUM = '#,##0'        # 千分位数字


# ==============================
# 1. 生成样例数据
# ==============================
def generate_sample_data(n=200):
    random.seed(42); np.random.seed(42)
    products = {
        "无线蓝牙耳机": {"price": (89, 199), "cost_ratio": 0.45},
        "智能手表": {"price": (299, 599), "cost_ratio": 0.40},
        "便携充电宝": {"price": (49, 129), "cost_ratio": 0.50},
        "USB-C扩展坞": {"price": (79, 189), "cost_ratio": 0.38},
        "机械键盘": {"price": (149, 349), "cost_ratio": 0.42},
        "手机壳套装": {"price": (29, 69), "cost_ratio": 0.30},
    }
    regions = ["华东", "华南", "华北", "西南", "华中"]
    channels = ["抖音", "淘宝", "拼多多", "京东", "私域"]
    customer_types = ["新客", "老客", "VIP"]
    start_date = datetime(2026, 1, 1)
    data = []
    for i in range(n):
        product = random.choice(list(products.keys()))
        info = products[product]
        price = round(random.uniform(*info["price"]), 2)
        quantity = random.randint(1, 50)
        cost = round(price * quantity * info["cost_ratio"], 2)
        revenue = round(price * quantity, 2)
        profit = round(revenue - cost, 2)
        date = start_date + timedelta(days=random.randint(0, 90))
        data.append({
            "日期": date.strftime("%Y-%m-%d"), "产品": product,
            "区域": random.choice(regions), "渠道": random.choice(channels),
            "客户类型": random.choice(customer_types), "单价": price,
            "数量": quantity, "销售额": revenue, "成本": cost, "利润": profit,
        })
    df = pd.DataFrame(data)
    df["日期"] = pd.to_datetime(df["日期"])
    df["月份"] = df["日期"].dt.to_period("M").astype(str)
    df = df.sort_values("日期").reset_index(drop=True)
    return df


# ==============================
# 2. 工具函数
# ==============================
def set_col_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def write_header_row(ws, row, headers, col_start=1):
    for j, h in enumerate(headers, col_start):
        cell = ws.cell(row=row, column=j, value=h)
        cell.font = FONT_HEADER
        cell.fill = FILL_GREEN_DARK
        cell.alignment = ALIGN_CENTER
        cell.border = BORDER_THIN


def write_cell(ws, row, col, value, fmt=None, font=FONT_BODY, fill=None, align=ALIGN_CENTER):
    """写入单元格，可选数字格式"""
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = font
    cell.alignment = align
    cell.border = BORDER_THIN
    if fill:
        cell.fill = fill
    if fmt:
        cell.number_format = fmt
    return cell


# ==============================
# 3. 创建报表
# ==============================
def create_report(df, output_path="销售数据自动化报表.xlsx"):
    # 确保有月份列
    if "日期" in df.columns and "月份" not in df.columns:
        df = df.copy()
        df["日期"] = pd.to_datetime(df["日期"])
        df["月份"] = df["日期"].dt.to_period("M").astype(str)
    wb = Workbook()

    # ---------- Sheet 1: 数据总览 ----------
    ws1 = wb.active
    ws1.title = "📊 数据总览"
    ws1.sheet_properties.tabColor = COLORS["green"]

    ws1.merge_cells("B2:H2")
    ws1.cell(row=2, column=2, value="销售数据自动化报表").font = FONT_TITLE
    ws1.cell(row=2, column=2).alignment = ALIGN_LEFT
    ws1.merge_cells("B3:H3")
    report_date = datetime.now().strftime("%Y年%m月%d日")
    ws1.cell(row=3, column=2, value=f"自动生成 · {report_date} · 林昕AI").font = FONT_MUTED

    # KPI cards
    total_revenue = df["销售额"].sum()
    total_profit = df["利润"].sum()
    total_orders = len(df)
    avg_order = df["销售额"].mean()
    profit_margin = (total_profit / total_revenue) if total_revenue > 0 else 0

    kpis = [
        ("总销售额", total_revenue, FMT_CNY),
        ("总利润", total_profit, FMT_CNY),
        ("订单数", total_orders, FMT_NUM),
        ("客单价", avg_order, FMT_CNY),
        ("利润率", profit_margin, FMT_PCT),
    ]
    kpi_row = 5
    for i, (label, value, fmt) in enumerate(kpis):
        col = 2 + i * 2
        for r in range(kpi_row, kpi_row + 3):
            for c in range(col, col + 2):
                cell = ws1.cell(row=r, column=c)
                cell.fill = FILL_GREEN_BG if i % 2 == 0 else FILL_WHITE
                cell.border = BORDER_THIN
        ws1.merge_cells(start_row=kpi_row, start_column=col, end_row=kpi_row, end_column=col + 1)
        ws1.merge_cells(start_row=kpi_row + 1, start_column=col, end_row=kpi_row + 1, end_column=col + 1)
        write_cell(ws1, kpi_row + 1, col, value, fmt=fmt, font=FONT_METRIC)
        ws1.merge_cells(start_row=kpi_row + 2, start_column=col, end_row=kpi_row + 2, end_column=col + 1)
        write_cell(ws1, kpi_row + 2, col, label, font=FONT_METRIC_LABEL)

    # 月度趋势表 — 写入数字而非字符串
    monthly = df.groupby("月份").agg(
        销售额=("销售额", "sum"), 利润=("利润", "sum"), 订单数=("销售额", "count"),
    ).reset_index()

    trend_start = 10
    ws1.merge_cells(f"B{trend_start}:H{trend_start}")
    ws1.cell(row=trend_start, column=2, value="📈 月度趋势").font = FONT_BODY_BOLD

    write_header_row(ws1, trend_start + 1, ["月份", "销售额", "利润", "订单数", "客单价", "利润率"], 2)
    for idx, row_data in monthly.iterrows():
        r = trend_start + 2 + idx
        avg = row_data["销售额"] / row_data["订单数"] if row_data["订单数"] > 0 else 0
        margin = row_data["利润"] / row_data["销售额"] if row_data["销售额"] > 0 else 0
        fill = FILL_SURFACE if idx % 2 == 0 else FILL_WHITE
        write_cell(ws1, r, 2, row_data["月份"], fill=fill)
        write_cell(ws1, r, 3, row_data["销售额"], fmt=FMT_CNY, fill=fill)
        write_cell(ws1, r, 4, row_data["利润"], fmt=FMT_CNY, fill=fill)
        write_cell(ws1, r, 5, row_data["订单数"], fmt=FMT_NUM, fill=fill)
        write_cell(ws1, r, 6, avg, fmt=FMT_CNY, fill=fill)
        write_cell(ws1, r, 7, margin, fmt=FMT_PCT, fill=fill)

    # 折线图 — 月度销售额
    n_months = len(monthly)
    chart_row = trend_start + 2 + n_months + 2
    chart1 = LineChart()
    chart1.title = "月度销售额趋势"
    chart1.style = 10
    chart1.width = 22; chart1.height = 12
    chart1.y_axis.title = "销售额 (元)"

    data_ref = Reference(ws1, min_col=3, min_row=trend_start + 1,
                         max_col=4, max_row=trend_start + 1 + n_months)
    cats_ref = Reference(ws1, min_col=2, min_row=trend_start + 2,
                         max_row=trend_start + 1 + n_months)
    chart1.add_data(data_ref, titles_from_data=True)
    chart1.set_categories(cats_ref)
    chart1.series[0].graphicalProperties.line.solidFill = COLORS["green"]
    chart1.series[0].graphicalProperties.line.width = 28000
    chart1.series[1].graphicalProperties.line.solidFill = COLORS["orange"]
    ws1.add_chart(chart1, f"B{chart_row}")

    # 产品占比饼图 — 写入数字
    product_sales = df.groupby("产品")["销售额"].sum().reset_index()
    temp_col = 10
    ws1.cell(row=trend_start + 1, column=temp_col, value="产品").font = FONT_MUTED
    ws1.cell(row=trend_start + 1, column=temp_col + 1, value="销售额").font = FONT_MUTED
    for idx, prow in product_sales.iterrows():
        ws1.cell(row=trend_start + 2 + idx, column=temp_col, value=prow["产品"])
        write_cell(ws1, trend_start + 2 + idx, temp_col + 1, prow["销售额"], fmt=FMT_CNY)

    pie1 = PieChart()
    pie1.title = "产品销售额占比"
    pie1.style = 10; pie1.width = 16; pie1.height = 12
    pie_data = Reference(ws1, min_col=temp_col + 1, min_row=trend_start + 1,
                         max_row=trend_start + 1 + len(product_sales))
    pie_cats = Reference(ws1, min_col=temp_col, min_row=trend_start + 2,
                         max_row=trend_start + 1 + len(product_sales))
    pie1.add_data(pie_data, titles_from_data=True)
    pie1.set_categories(pie_cats)
    pie1.dataLabels = DataLabelList()
    pie1.dataLabels.showPercent = True
    pie1.dataLabels.showCatName = True
    green_shades = ["22C55E", "16A34A", "15803D", "166534", "4ADE80", "86EFAC"]
    for i in range(len(product_sales)):
        pt = DataPoint(idx=i)
        pt.graphicalProperties.solidFill = green_shades[i % len(green_shades)]
        pie1.series[0].data_points.append(pt)
    ws1.add_chart(pie1, f"L{chart_row}")

    set_col_widths(ws1, [2, 16, 18, 18, 14, 14, 16, 16, 2, 14, 18, 18])

    # ---------- Sheet 2: 销售明细 ----------
    ws2 = wb.create_sheet("📋 销售明细")
    ws2.sheet_properties.tabColor = "16A34A"

    ws2.merge_cells("A1:K1")
    ws2.cell(row=1, column=1, value="销售明细数据").font = FONT_TITLE

    headers = ["日期", "产品", "区域", "渠道", "客户类型", "单价", "数量", "销售额", "成本", "利润", "利润率"]
    write_header_row(ws2, 3, headers)

    detail_df = df.sort_values("日期", ascending=False).reset_index(drop=True)
    for idx, row_data in detail_df.iterrows():
        r = 4 + idx
        margin = (row_data["利润"] / row_data["销售额"]) if row_data["销售额"] > 0 else 0
        fill = FILL_RED_LIGHT if margin < 0.3 else (FILL_GREEN_BG if margin > 0.6 else (FILL_SURFACE if idx % 2 == 0 else FILL_WHITE))
        write_cell(ws2, r, 1, row_data["日期"].strftime("%Y-%m-%d"), fill=fill)
        write_cell(ws2, r, 2, row_data["产品"], fill=fill)
        write_cell(ws2, r, 3, row_data["区域"], fill=fill)
        write_cell(ws2, r, 4, row_data["渠道"], fill=fill)
        write_cell(ws2, r, 5, row_data["客户类型"], fill=fill)
        write_cell(ws2, r, 6, row_data["单价"], fmt=FMT_CNY, fill=fill, align=ALIGN_RIGHT)
        write_cell(ws2, r, 7, row_data["数量"], fmt=FMT_NUM, fill=fill, align=ALIGN_RIGHT)
        write_cell(ws2, r, 8, row_data["销售额"], fmt=FMT_CNY, fill=fill, align=ALIGN_RIGHT)
        write_cell(ws2, r, 9, row_data["成本"], fmt=FMT_CNY, fill=fill, align=ALIGN_RIGHT)
        write_cell(ws2, r, 10, row_data["利润"], fmt=FMT_CNY, fill=fill, align=ALIGN_RIGHT)
        write_cell(ws2, r, 11, margin, fmt=FMT_PCT, fill=fill)

    set_col_widths(ws2, [14, 18, 10, 12, 12, 12, 10, 16, 16, 16, 12])

    # ---------- Sheet 3: 区域分析 ----------
    ws3 = wb.create_sheet("🗺️ 区域分析")
    ws3.sheet_properties.tabColor = "15803D"

    ws3.merge_cells("A1:H1")
    ws3.cell(row=1, column=1, value="区域销售分析").font = FONT_TITLE

    region_data = df.groupby("区域").agg(
        销售额=("销售额", "sum"), 利润=("利润", "sum"),
        订单数=("销售额", "count"), 客单价=("销售额", "mean"),
    ).reset_index().sort_values("销售额", ascending=False)

    total_rev = region_data["销售额"].sum()
    write_header_row(ws3, 3, ["区域", "销售额", "利润", "订单数", "客单价", "利润率", "占比"])
    for idx, row_data in region_data.iterrows():
        r = 4 + idx
        margin = row_data["利润"] / row_data["销售额"] if row_data["销售额"] > 0 else 0
        share = row_data["销售额"] / total_rev if total_rev > 0 else 0
        fill = FILL_GREEN_BG if idx == 0 else (FILL_SURFACE if idx % 2 == 0 else FILL_WHITE)
        write_cell(ws3, r, 1, row_data["区域"], fill=fill)
        write_cell(ws3, r, 2, row_data["销售额"], fmt=FMT_CNY, fill=fill, align=ALIGN_RIGHT)
        write_cell(ws3, r, 3, row_data["利润"], fmt=FMT_CNY, fill=fill, align=ALIGN_RIGHT)
        write_cell(ws3, r, 4, row_data["订单数"], fmt=FMT_NUM, fill=fill, align=ALIGN_RIGHT)
        write_cell(ws3, r, 5, row_data["客单价"], fmt=FMT_CNY, fill=fill, align=ALIGN_RIGHT)
        write_cell(ws3, r, 6, margin, fmt=FMT_PCT, fill=fill)
        write_cell(ws3, r, 7, share, fmt=FMT_PCT, fill=fill)

    # 区域柱状图
    n_regions = len(region_data)
    chart_row3 = 4 + n_regions + 2
    bar1 = BarChart()
    bar1.type = "col"; bar1.title = "各区域销售额对比"; bar1.style = 10
    bar1.width = 20; bar1.height = 12
    bar_data = Reference(ws3, min_col=2, min_row=3, max_col=3, max_row=3 + n_regions)
    bar_cats = Reference(ws3, min_col=1, min_row=4, max_row=3 + n_regions)
    bar1.add_data(bar_data, titles_from_data=True)
    bar1.set_categories(bar_cats)
    bar1.series[0].graphicalProperties.solidFill = COLORS["green"]
    bar1.series[1].graphicalProperties.solidFill = COLORS["orange"]
    ws3.add_chart(bar1, f"A{chart_row3}")

    set_col_widths(ws3, [12, 16, 16, 10, 14, 10, 10])

    # ---------- Sheet 4: 产品分析 ----------
    ws4 = wb.create_sheet("📦 产品分析")
    ws4.sheet_properties.tabColor = "4ADE80"

    ws4.merge_cells("A1:H1")
    ws4.cell(row=1, column=1, value="产品销售分析").font = FONT_TITLE

    prod_data = df.groupby("产品").agg(
        销售额=("销售额", "sum"), 利润=("利润", "sum"),
        订单数=("销售额", "count"), 平均单价=("单价", "mean"),
    ).reset_index().sort_values("销售额", ascending=False)

    write_header_row(ws4, 3, ["产品", "销售额", "利润", "订单数", "平均单价", "利润率", "排名"])
    for idx, row_data in prod_data.iterrows():
        r = 4 + idx
        margin = row_data["利润"] / row_data["销售额"] if row_data["销售额"] > 0 else 0
        fill = FILL_GREEN_BG if idx == 0 else (FILL_SURFACE if idx % 2 == 0 else FILL_WHITE)
        write_cell(ws4, r, 1, row_data["产品"], fill=fill)
        write_cell(ws4, r, 2, row_data["销售额"], fmt=FMT_CNY, fill=fill, align=ALIGN_RIGHT)
        write_cell(ws4, r, 3, row_data["利润"], fmt=FMT_CNY, fill=fill, align=ALIGN_RIGHT)
        write_cell(ws4, r, 4, row_data["订单数"], fmt=FMT_NUM, fill=fill, align=ALIGN_RIGHT)
        write_cell(ws4, r, 5, row_data["平均单价"], fmt=FMT_CNY, fill=fill, align=ALIGN_RIGHT)
        write_cell(ws4, r, 6, margin, fmt=FMT_PCT, fill=fill)
        write_cell(ws4, r, 7, f"#{idx + 1}", fill=fill)

    n_prods = len(prod_data)
    chart_row4 = 4 + n_prods + 2
    bar2 = BarChart()
    bar2.type = "col"; bar2.title = "各产品销售额对比"; bar2.style = 10
    bar2.width = 20; bar2.height = 12
    bar2_data = Reference(ws4, min_col=2, min_row=3, max_col=3, max_row=3 + n_prods)
    bar2_cats = Reference(ws4, min_col=1, min_row=4, max_row=3 + n_prods)
    bar2.add_data(bar2_data, titles_from_data=True)
    bar2.set_categories(bar2_cats)
    bar2.series[0].graphicalProperties.solidFill = COLORS["green"]
    bar2.series[1].graphicalProperties.solidFill = COLORS["orange"]
    ws4.add_chart(bar2, f"A{chart_row4}")

    set_col_widths(ws4, [18, 16, 16, 10, 14, 10, 10])

    # ---------- Sheet 5: 渠道分析 ----------
    ws5 = wb.create_sheet("📱 渠道分析")
    ws5.sheet_properties.tabColor = "22C55E"

    ws5.merge_cells("A1:H1")
    ws5.cell(row=1, column=1, value="渠道销售分析").font = FONT_TITLE

    channel_data = df.groupby("渠道").agg(
        销售额=("销售额", "sum"), 利润=("利润", "sum"), 订单数=("销售额", "count"),
    ).reset_index().sort_values("销售额", ascending=False)

    write_header_row(ws5, 3, ["渠道", "销售额", "利润", "订单数", "客单价", "利润率", "占比"])
    for idx, row_data in channel_data.iterrows():
        r = 4 + idx
        margin = row_data["利润"] / row_data["销售额"] if row_data["销售额"] > 0 else 0
        share = row_data["销售额"] / total_rev if total_rev > 0 else 0
        avg = row_data["销售额"] / row_data["订单数"] if row_data["订单数"] > 0 else 0
        fill = FILL_GREEN_BG if idx == 0 else (FILL_SURFACE if idx % 2 == 0 else FILL_WHITE)
        write_cell(ws5, r, 1, row_data["渠道"], fill=fill)
        write_cell(ws5, r, 2, row_data["销售额"], fmt=FMT_CNY, fill=fill, align=ALIGN_RIGHT)
        write_cell(ws5, r, 3, row_data["利润"], fmt=FMT_CNY, fill=fill, align=ALIGN_RIGHT)
        write_cell(ws5, r, 4, row_data["订单数"], fmt=FMT_NUM, fill=fill, align=ALIGN_RIGHT)
        write_cell(ws5, r, 5, avg, fmt=FMT_CNY, fill=fill, align=ALIGN_RIGHT)
        write_cell(ws5, r, 6, margin, fmt=FMT_PCT, fill=fill)
        write_cell(ws5, r, 7, share, fmt=FMT_PCT, fill=fill)

    # 渠道柱状图
    n_channels = len(channel_data)
    chart_row5 = 4 + n_channels + 2
    bar3 = BarChart()
    bar3.type = "col"; bar3.title = "各渠道销售额对比"; bar3.style = 10
    bar3.width = 20; bar3.height = 12
    bar3_data = Reference(ws5, min_col=2, min_row=3, max_col=3, max_row=3 + n_channels)
    bar3_cats = Reference(ws5, min_col=1, min_row=4, max_row=3 + n_channels)
    bar3.add_data(bar3_data, titles_from_data=True)
    bar3.set_categories(bar3_cats)
    bar3.series[0].graphicalProperties.solidFill = COLORS["green"]
    bar3.series[1].graphicalProperties.solidFill = COLORS["orange"]
    ws5.add_chart(bar3, f"A{chart_row5}")

    set_col_widths(ws5, [12, 16, 16, 10, 12, 10, 10])

    # 保存
    wb.save(output_path)
    return output_path


# ==============================
# Main
# ==============================
if __name__ == "__main__":
    print("🔄 正在生成销售数据...")
    if len(sys.argv) > 1 and os.path.exists(sys.argv[1]):
        print(f"📂 读取数据文件: {sys.argv[1]}")
        df = pd.read_csv(sys.argv[1])
    else:
        print("📊 使用内置样例数据 (200条模拟销售记录)")
        df = generate_sample_data(200)

    print(f"   数据量: {len(df)} 条记录")
    print(f"   时间范围: {df['日期'].min()} ~ {df['日期'].max()}")
    print(f"   产品种类: {df['产品'].nunique()} 种")
    print(f"   覆盖区域: {df['区域'].nunique()} 个")

    output = create_report(df)
    print(f"\n✅ 报表已生成: {output}")
    print(f"   文件大小: {os.path.getsize(output) / 1024:.1f} KB")
    print(f"   包含 5 个工作表: 数据总览 / 销售明细 / 区域分析 / 产品分析 / 渠道分析")
